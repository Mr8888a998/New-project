from __future__ import annotations

import csv
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from handicap_ai.adapters.football_data import FootballDataCsvAdapter
from handicap_ai.database import Database
from handicap_ai.ingest import ingest_bundles


SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xlsm"}
DEFAULT_COMPETITION = "HIST"


@dataclass(frozen=True)
class ImportSummary:
    files_imported: int
    files_skipped: int
    matches_imported: int
    errors: tuple[str, ...]


def import_history_folder(db: Database, folder: Path, season: str) -> ImportSummary:
    files_imported = 0
    files_skipped = 0
    matches_imported = 0
    errors: list[str] = []

    for child in sorted(Path(folder).iterdir()):
        if child.suffix.lower() not in SUPPORTED_SUFFIXES:
            files_skipped += 1
            continue

        try:
            matches_imported += _import_history_file(db, child, season)
        except Exception as exc:
            files_skipped += 1
            errors.append(f"{child.name}: {exc}")
        else:
            files_imported += 1

    return ImportSummary(
        files_imported=files_imported,
        files_skipped=files_skipped,
        matches_imported=matches_imported,
        errors=tuple(errors),
    )


def _import_history_file(db: Database, path: Path, season: str) -> int:
    if path.suffix.lower() == ".csv":
        prepared_path = _prepare_csv(path)
    else:
        prepared_path = _excel_to_csv(path)

    try:
        bundles = FootballDataCsvAdapter(prepared_path, season=season).load()
        return ingest_bundles(db, bundles)
    finally:
        prepared_path.unlink(missing_ok=True)


def _prepare_csv(path: Path) -> Path:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return _write_prepared_csv(csv.reader(handle))


def _excel_to_csv(path: Path) -> Path:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        return _write_prepared_csv(rows)
    finally:
        workbook.close()


def _write_prepared_csv(
    rows: Iterable[Sequence[Any]],
) -> Path:
    iterator = iter(rows)
    try:
        raw_header = next(iterator)
    except StopIteration as exc:
        raise ValueError("history file is empty") from exc

    header = [_clean_cell(cell) for cell in raw_header]
    output_header = _prepared_header(header)
    temp_path = _temporary_csv_path()

    try:
        with temp_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(output_header)
            for row in iterator:
                writer.writerow(_prepared_row(row, header))
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise

    return temp_path


def _prepared_header(header: list[str]) -> list[str]:
    prepared = list(header)
    if "Div" not in prepared:
        prepared.insert(0, "Div")
    if "BbAv>2.5" not in prepared and "B365>2.5" in header:
        prepared.append("BbAv>2.5")
    if "BbAv<2.5" not in prepared and "B365<2.5" in header:
        prepared.append("BbAv<2.5")
    return prepared


def _prepared_row(row: Sequence[Any], header: list[str]) -> list[str]:
    values = [_clean_cell(cell) for cell in row]
    prepared = list(values)
    if "Div" not in header:
        prepared.insert(0, DEFAULT_COMPETITION)
    if "BbAv>2.5" not in header and "B365>2.5" in header:
        prepared.append(_value_for_header(values, header, "B365>2.5"))
    if "BbAv<2.5" not in header and "B365<2.5" in header:
        prepared.append(_value_for_header(values, header, "B365<2.5"))
    return prepared


def _value_for_header(values: Sequence[str], header: list[str], field: str) -> str:
    index = header.index(field)
    if index >= len(values):
        return ""
    return values[index]


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _temporary_csv_path() -> Path:
    handle = NamedTemporaryFile(
        mode="w",
        newline="",
        encoding="utf-8",
        suffix=".csv",
        delete=False,
    )
    handle.close()
    return Path(handle.name)
