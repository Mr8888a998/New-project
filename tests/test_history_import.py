from pathlib import Path

from handicap_ai.database import Database
from handicap_ai.history_import import import_history_folder


def test_import_history_folder_imports_supported_csv(tmp_path):
    db = Database(tmp_path / "handicap.sqlite")
    db.migrate()

    summary = import_history_folder(
        db=db,
        folder=Path("tests/fixtures/history_folder"),
        season="2026",
    )

    assert summary.files_imported == 1
    assert summary.files_skipped == 0
    assert summary.matches_imported == 1
    assert db.find_matches_by_names("England", "Panama")


def test_import_history_folder_is_idempotent(tmp_path):
    db = Database(tmp_path / "handicap.sqlite")
    db.migrate()

    import_history_folder(db=db, folder=Path("tests/fixtures/history_folder"), season="2026")
    import_history_folder(db=db, folder=Path("tests/fixtures/history_folder"), season="2026")

    rows = db.find_matches_by_names("England", "Panama")
    assert len(rows) == 1


def test_import_history_folder_imports_xlsx(tmp_path):
    from openpyxl import Workbook

    folder = tmp_path / "history"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Date",
            "HomeTeam",
            "AwayTeam",
            "FTHG",
            "FTAG",
            "B365H",
            "B365D",
            "B365A",
            "AHh",
            "B365AHH",
            "B365AHA",
            "B365>2.5",
            "B365<2.5",
        ]
    )
    sheet.append(
        [
            "02/01/26",
            "Portugal",
            "Uzbekistan",
            3,
            1,
            1.25,
            5.80,
            11.00,
            -1.75,
            1.91,
            1.96,
            1.88,
            1.98,
        ]
    )
    workbook.save(folder / "sample.xlsx")
    db = Database(tmp_path / "handicap.sqlite")
    db.migrate()

    summary = import_history_folder(db=db, folder=folder, season="2026")

    assert summary.files_imported == 1
    assert summary.matches_imported == 1
    assert db.find_matches_by_names("Portugal", "Uzbekistan")


def test_import_history_folder_imports_xlsx_native_date_cell(tmp_path):
    from datetime import datetime

    from openpyxl import Workbook

    folder = tmp_path / "history"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Date",
            "HomeTeam",
            "AwayTeam",
            "FTHG",
            "FTAG",
            "B365H",
            "B365D",
            "B365A",
            "AHh",
            "B365AHH",
            "B365AHA",
            "B365>2.5",
            "B365<2.5",
        ]
    )
    sheet.append(
        [
            datetime(2026, 1, 2),
            "France",
            "Morocco",
            2,
            1,
            1.42,
            4.60,
            7.50,
            -1.25,
            1.92,
            1.95,
            1.87,
            1.99,
        ]
    )
    workbook.save(folder / "native-date.xlsx")
    db = Database(tmp_path / "handicap.sqlite")
    db.migrate()

    summary = import_history_folder(db=db, folder=folder, season="2026")

    assert summary.errors == ()
    assert summary.files_imported == 1
    assert summary.matches_imported == 1
    assert db.find_matches_by_names("France", "Morocco")


def test_import_history_folder_imports_xlsm(tmp_path):
    from openpyxl import Workbook

    folder = tmp_path / "history"
    folder.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Date",
            "HomeTeam",
            "AwayTeam",
            "FTHG",
            "FTAG",
            "B365H",
            "B365D",
            "B365A",
            "AHh",
            "B365AHH",
            "B365AHA",
            "B365>2.5",
            "B365<2.5",
        ]
    )
    sheet.append(
        [
            "03/01/26",
            "Croatia",
            "Panama",
            2,
            0,
            1.38,
            4.90,
            8.75,
            -1.25,
            1.90,
            1.97,
            1.86,
            2.00,
        ]
    )
    workbook.save(folder / "sample.xlsm")
    db = Database(tmp_path / "handicap.sqlite")
    db.migrate()

    summary = import_history_folder(db=db, folder=folder, season="2026")

    assert summary.errors == ()
    assert summary.files_imported == 1
    assert summary.matches_imported == 1
    assert db.find_matches_by_names("Croatia", "Panama")
